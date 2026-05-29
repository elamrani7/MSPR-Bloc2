"""
Gantt Excel — MSPR TPRE921 COFRAP  |  Style professionnel moderne
python3 scripts/generate_gantt.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "../docs/planning/Gantt_MSPR_COFRAP.xlsx")

# ── Palette moderne ──────────────────────────────────────────────────────────
P = {
    # Headers
    "h1":      "0D1B2A",   # bleu nuit (titre)
    "h2":      "1B3A5C",   # bleu foncé (sous-titre)
    "h_day":   "1565C0",   # bleu jour
    "h_hour":  "E3F2FD",   # bleu très clair (heure)
    "h_hour_t":"1565C0",
    # Assignees  (bar_dark, bar_light, text_on_light)
    "youssef":  ("1565C0","BBDEFB","0D47A1"),
    "soulaiman":("2E7D32","C8E6C9","1B5E20"),
    "hamza":    ("E65100","FFE0B2","BF360C"),
    "p4":       ("6A1B9A","E1BEE7","4A148C"),
    "all":      ("37474F","ECEFF1","263238"),
    "milestone":("B71C1C","FFCDD2","B71C1C"),
    # Rows
    "row_odd":  "FAFCFF",
    "row_even": "F0F4FA",
    "section":  "E8F0FE",
    "sec_txt":  "1A237E",
    "sep":      "CFD8DC",
    # Borders
    "brd":      "CFD8DC",
    "brd_day":  "90CAF9",
    "crit":     "C62828",
    "white":    "FFFFFF",
}

def fl(h): return PatternFill("solid", fgColor=h)
def fn(bold=False, color="212121", sz=9, italic=False):
    return Font(bold=bold, color=color, size=sz, italic=italic, name="Calibri")
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(color=None, thick=False, left_c=None):
    c = color or P["brd"]
    st = "medium" if thick else "thin"
    def S(col=None, style=None):
        return Side(border_style=style or st, color=col or c)
    return Border(
        left  = S(col=left_c, style="medium") if left_c else S(),
        right = S(), top=S(), bottom=S()
    )
def side(c, st="thin"): return Side(border_style=st, color=c)

# ── Tâches ───────────────────────────────────────────────────────────────────
TASKS = [
    # J1
    dict(id="MACA-73",sec="  J1  ·  Lundi 20/04/2026  ·  Cadrage & Organisation",
         title="Initialisation dépôt Git & structure projet",
         who="youssef", day=0,sh=9, dh=1,crit=False),
    dict(id="MACA-74",sec=None,
         title="Configuration Jira — Epics, Sprints, Tickets MACA",
         who="youssef", day=0,sh=10,dh=1,crit=False),
    dict(id="MACA-75",sec=None,
         title="Rédaction Gantt & planning projet (4 jours)",
         who="youssef", day=0,sh=11,dh=2,crit=False),
    dict(id="MACA-76",sec=None,
         title="Architecture technique COFRAP  (Frontend ↔ OpenFaaS ↔ DB)",
         who="all",     day=0,sh=13,dh=3,crit=True),
    # J2
    dict(id="MACA-77",sec="  J2  ·  Mardi 21/04/2026  ·  Infrastructure & Développement",
         title="Setup environnement dev local (.env · dépendances Python)",
         who="all",      day=1,sh=9, dh=2,crit=True),
    dict(id="MACA-78",sec=None,
         title="Schéma PostgreSQL — table users (password_hash, totp_secret)",
         who="soulaiman",day=1,sh=9, dh=2,crit=True),
    dict(id="MACA-79",sec=None,
         title="Installation K3S + kubectl  (control-plane + worker node)",
         who="hamza",    day=1,sh=11,dh=3,crit=True),
    dict(id="MACA-80",sec=None,
         title="Déploiement OpenFaaS Community via Helm  (faas-netes)",
         who="hamza",    day=1,sh=14,dh=2,crit=True),
    dict(id="MACA-81",sec=None,
         title="Déploiement PostgreSQL Helm + secrets OpenFaaS",
         who="soulaiman",day=1,sh=11,dh=2,crit=True),
    dict(id="MACA-82",sec=None,
         title="Développement fonction generate-password  (24 chars + bcrypt + QR Code)",
         who="soulaiman",day=1,sh=13,dh=4,crit=True),
    # J3
    dict(id="MACA-83",sec="  J3  ·  Mercredi 22/04/2026  ·  Intégration & Frontend",
         title="Développement fonction generate-2fa  (TOTP secret + QR Code base64)",
         who="soulaiman",day=2,sh=9, dh=3,crit=True),
    dict(id="MACA-84",sec=None,
         title="Développement fonction authenticate  (password + TOTP + expiry 6 mois)",
         who="hamza",    day=2,sh=12,dh=3,crit=True),
    dict(id="MACA-85",sec=None,
         title="Frontend React — Pages Create Account + génération QR Codes",
         who="p4",       day=2,sh=9, dh=4,crit=False),
    dict(id="MACA-86",sec=None,
         title="Frontend React — Pages Login + Renouvellement credentials",
         who="p4",       day=2,sh=13,dh=3,crit=False),
    dict(id="MACA-87",sec=None,
         title="Build & push images Docker Hub  (3 fonctions OpenFaaS)",
         who="hamza",    day=2,sh=9, dh=2,crit=True),
    dict(id="MACA-88",sec=None,
         title="Intégration Frontend ↔ OpenFaaS Gateway + tests Postman",
         who="all",      day=2,sh=15,dh=2,crit=True),
    # J4
    dict(id="MACA-89",sec="  J4  ·  Jeudi 23/04/2026  ·  Tests & Soutenance",
         title="Tests E2E complets — 4 scénarios Postman",
         who="all",      day=3,sh=8, dh=2,crit=True),
    dict(id="MACA-90",sec=None,
         title="Rédaction dossier technique final (architecture, screenshots, code)",
         who="p4",       day=3,sh=8, dh=3,crit=False),
    dict(id="MACA-91",sec=None,
         title="Déploiement Frontend sur K3S  (Ingress Traefik + CORS)",
         who="hamza",    day=3,sh=8, dh=2,crit=True),
    dict(id="MACA-92",sec=None,
         title="Préparation support soutenance  (slides + démo live)",
         who="youssef",  day=3,sh=10,dh=2,crit=True),
    dict(id="MACA-93",sec=None,
         title="Répétition générale soutenance  (20 min chrono) + relecture dossier",
         who="all",      day=3,sh=12,dh=2,crit=True),
    dict(id="🎯",sec=None,
         title="SOUTENANCE MSPR  —  Jeudi 23/04/2026  à  14h00",
         who="milestone",day=3,sh=14,dh=0,crit=True),
]

DAYS = [
    dict(label="J1  ·  Lundi 20/04",   bh=9, nh=9),
    dict(label="J2  ·  Mardi 21/04",   bh=9, nh=9),
    dict(label="J3  ·  Mercredi 22/04",bh=9, nh=9),
    dict(label="J4  ·  Jeudi 23/04",   bh=8, nh=7),
]
GAP=1; INFO=6   # cols: strip|id|title|who|dur|crit

def day_off(d): return sum(DAYS[i]["nh"]+GAP for i in range(d))
def slots(t):
    d=t["day"]; cfg=DAYS[d]; base=day_off(d); s=t["sh"]-cfg["bh"]
    if t["dh"]==0: return [base+s]
    return [base+s+i for i in range(t["dh"]) if s+i<cfg["nh"]]

TL = sum(c["nh"] for c in DAYS)+GAP*(len(DAYS)-1)
TC = INFO+TL

def build():
    wb=Workbook(); ws=wb.active; ws.title="Gantt MSPR COFRAP"

    # largeurs colonnes
    ws.column_dimensions["A"].width = 1.2   # strip couleur
    ws.column_dimensions["B"].width = 9.5   # ID
    ws.column_dimensions["C"].width = 44    # Tâche
    ws.column_dimensions["D"].width = 11    # Responsable
    ws.column_dimensions["E"].width = 5.5   # Dur
    ws.column_dimensions["F"].width = 4     # Crit
    sl=0
    for d,cfg in enumerate(DAYS):
        for _ in range(cfg["nh"]):
            ws.column_dimensions[get_column_letter(INFO+1+sl)].width=3.4; sl+=1
        if d<len(DAYS)-1:
            ws.column_dimensions[get_column_letter(INFO+1+sl)].width=0.8; sl+=1

    # ── ROW 1 : Titre ─────────────────────────────────────────────────────
    ws.row_dimensions[1].height=32
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=TC)
    c=ws.cell(1,1,"  MSPR TPRE921  ·  COFRAP Serverless Auth System  ·  Diagramme de Gantt  ·  20/04/2026 → 23/04/2026  ·  Soutenance J4 · 14h00")
    c.fill=fl(P["h1"]); c.font=fn(True,"FFFFFF",13); c.alignment=aln("left")
    c.border=Border(bottom=side("1565C0","medium"))

    # ── ROW 2 : Sous-titre ─────────────────────────────────────────────────
    ws.row_dimensions[2].height=15
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=TC)
    c=ws.cell(2,1,"     Python  ·  OpenFaaS Community  ·  K3S  ·  PostgreSQL  ·  React.js  ·  Helm  ·  Docker Hub     |     Équipe : Youssef  ·  Soulaiman  ·  Hamza  ·  [PRÉNOM4]     |     🔴 = Chemin critique")
    c.fill=fl(P["h2"]); c.font=fn(False,"BBDEFB",8,True); c.alignment=aln("left")

    # ── ROW 3 : Vide ───────────────────────────────────────────────────────
    ws.row_dimensions[3].height=4
    for ci in range(1,TC+1): ws.cell(3,ci).fill=fl("E8EEF7")

    # ── ROW 4 : En-têtes info + jours ─────────────────────────────────────
    ws.row_dimensions[4].height=22
    for ci,lbl,al_h in [(1,"","center"),(2,"ID","center"),(3,"  Tâche","left"),
                         (4,"Responsable","center"),(5,"Dur.","center"),(6,"","center")]:
        ws.merge_cells(start_row=4,start_column=ci,end_row=5,end_column=ci)
        c=ws.cell(4,ci,lbl)
        c.fill=fl(P["h1"]); c.font=fn(True,"FFFFFF",9)
        c.alignment=aln(al_h); c.border=Border(right=side("FFFFFF","thin"),bottom=side("1565C0"))

    for d,cfg in enumerate(DAYS):
        off=day_off(d); cs=INFO+1+off; ce=cs+cfg["nh"]-1
        ws.merge_cells(start_row=4,start_column=cs,end_row=4,end_column=ce)
        c=ws.cell(4,cs,cfg["label"])
        c.fill=fl(P["h_day"]); c.font=fn(True,"FFFFFF",11); c.alignment=aln()
        c.border=Border(left=side("FFFFFF","medium"),right=side("FFFFFF","medium"),
                        bottom=side("90CAF9","thin"))
        if d<len(DAYS)-1:
            gc=INFO+1+off+cfg["nh"]
            ws.merge_cells(start_row=4,start_column=gc,end_row=5,end_column=gc)
            ws.cell(4,gc).fill=fl(P["sep"])

    # ── ROW 5 : Heures ────────────────────────────────────────────────────
    ws.row_dimensions[5].height=13
    for d,cfg in enumerate(DAYS):
        off=day_off(d)
        for hi in range(cfg["nh"]):
            ca=INFO+1+off+hi
            c=ws.cell(5,ca,f"{cfg['bh']+hi}h")
            c.fill=fl(P["h_hour"]); c.font=fn(False,P["h_hour_t"],7)
            c.alignment=aln()
            c.border=Border(left=side(P["brd_day"]),right=side(P["brd_day"]),
                            bottom=side("1565C0","medium"))

    # ── TÂCHES ────────────────────────────────────────────────────────────
    row=6; last_sec=None
    WHO_LBL={"youssef":"Youssef","soulaiman":"Soulaiman","hamza":"Hamza",
              "p4":"[PRÉNOM4]","all":"Équipe","milestone":"ÉQUIPE"}

    for ti,t in enumerate(TASKS):
        # Section header
        if t.get("sec") and t["sec"]!=last_sec:
            last_sec=t["sec"]
            ws.row_dimensions[row].height=17
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=TC)
            c=ws.cell(row,1,t["sec"])
            c.fill=fl(P["section"]); c.font=fn(True,P["sec_txt"],10)
            c.alignment=aln("left")
            c.border=Border(left=side("1565C0","medium"),
                            bottom=side("1565C0","medium"),top=side("CFD8DC"))
            row+=1

        ws.row_dimensions[row].height=18
        ms=t["who"]=="milestone"
        bd_,lt_,_=P[t["who"]] if not ms else P["milestone"]
        crit=t["crit"]
        alt=P["row_odd"] if ti%2==0 else P["row_even"]
        active=set(slots(t))

        # Col A : bande couleur (strip)
        c=ws.cell(row,1)
        c.fill=fl(bd_)
        c.border=Border(right=side(lt_,"thin"))

        # Col B : ID
        c=ws.cell(row,2,t["id"])
        c.fill=fl("E8EEF7" if not ms else P["milestone"][1])
        c.font=fn(True,bd_,8)
        c.alignment=aln(); c.border=bdr(P["brd"])

        # Col C : Titre
        c=ws.cell(row,3,t["title"])
        c.fill=fl(P["milestone"][1] if ms else alt)
        c.font=fn(ms or crit, color=P["milestone"][0] if ms else ("212121"),sz=9)
        c.alignment=aln("left")
        c.border=Border(left=side(P["brd"]),right=side(P["brd"]),
                        top=side(P["brd"]),
                        bottom=side(P["crit"],"medium") if crit and not ms else side(P["brd"]))

        # Col D : Responsable
        c=ws.cell(row,4,WHO_LBL[t["who"]])
        c.fill=fl(bd_); c.font=fn(True,"FFFFFF",8); c.alignment=aln(); c.border=bdr(P["brd"])

        # Col E : Durée
        dur="◆" if ms else f"{t['dh']}h"
        c=ws.cell(row,5,dur)
        c.fill=fl(P["milestone"][1] if ms else alt)
        c.font=fn(ms,bd_ if ms else "546E7A",9); c.alignment=aln(); c.border=bdr(P["brd"])

        # Col F : Critique
        c=ws.cell(row,6,"●" if crit and not ms else "")
        c.fill=fl(alt)
        c.font=Font(bold=True,color=P["crit"] if crit else alt,size=10,name="Calibri")
        c.alignment=aln(); c.border=bdr(P["brd"])

        # Timeline
        sl_list=sorted(active)
        for d,cfg in enumerate(DAYS):
            off=day_off(d)
            for hi in range(cfg["nh"]):
                sl=off+hi; ca=INFO+1+sl; c=ws.cell(row,ca)
                if sl in active:
                    if ms:
                        c.value="◆"; c.fill=fl(P["milestone"][1])
                        c.font=Font(bold=True,color=P["milestone"][0],size=14,name="Calibri")
                        c.alignment=aln()
                        c.border=Border(left=side(P["milestone"][0],"medium"),
                                        right=side(P["milestone"][0],"medium"),
                                        top=side(P["milestone"][0],"medium"),
                                        bottom=side(P["milestone"][0],"medium"))
                    else:
                        c.fill=fl(bd_)
                        # bar caps
                        is_first = sl==min(active); is_last = sl==max(active)
                        c.border=Border(
                            left =side(P["crit"],"medium") if crit and is_first  else side(lt_,"medium" if is_first else "thin"),
                            right=side(P["crit"],"medium") if crit and is_last   else side(lt_,"medium" if is_last  else "thin"),
                            top  =side(P["crit"],"medium") if crit else side(lt_,"thin"),
                            bottom=side(P["crit"],"medium") if crit else side(lt_,"thin"),
                        )
                else:
                    c.fill=fl(alt)
                    c.border=Border(left=side(P["brd_day"] if d==0 and hi==0 else P["brd"]),
                                    right=side(P["brd"]),top=side(P["brd"]),bottom=side(P["brd"]))
            if d<len(DAYS)-1:
                ws.cell(row,INFO+1+off+cfg["nh"]).fill=fl(P["sep"])
        row+=1

    # ── LÉGENDE ───────────────────────────────────────────────────────────
    row+=1
    ws.row_dimensions[row].height=5
    for ci in range(1,TC+1): ws.cell(row,ci).fill=fl("E8EEF7")
    row+=1

    ws.row_dimensions[row].height=20
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=TC)
    c=ws.cell(row,1,"  LÉGENDE  —  Responsables & Conventions")
    c.fill=fl(P["h1"]); c.font=fn(True,"FFFFFF",10); c.alignment=aln("left"); row+=1

    LEG=[
        ("Youssef",    "youssef",  "Chef de Projet · Architecte Solution · DevOps"),
        ("Soulaiman",  "soulaiman","Développeur Backend Lead  (Python · OpenFaaS · PostgreSQL)"),
        ("Hamza",      "hamza",    "Développeur Backend Auth + DevOps K3S · Docker Hub"),
        ("[PRÉNOM4]",  "p4",       "Développeur Frontend React.js + Rédacteur dossier final"),
        ("Équipe",     "all",      "Tâche collective — toute l'équipe mobilisée"),
        ("● Crit.",    None,       "Chemin critique — retard = impact direct sur la soutenance"),
        ("◆ Jalon",    None,       "Milestone clé — Soutenance J4 à 14h00"),
    ]
    for li,(lbl,who,desc) in enumerate(LEG):
        ws.row_dimensions[row].height=16
        bc=P[who][0] if who else P["h_day"]
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
        c=ws.cell(row,1,f"  {lbl}"); c.fill=fl(bc); c.font=fn(True,"FFFFFF",9)
        c.alignment=aln("left"); c.border=Border(right=side("FFFFFF"),bottom=side(P["brd"]))
        ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=10)
        c=ws.cell(row,3,f"  {desc}"); c.fill=fl(P["row_odd"] if li%2==0 else P["row_even"])
        c.font=fn(sz=9,color="37474F"); c.alignment=aln("left"); c.border=bdr(P["brd"]); row+=1

    # ── Setup ──────────────────────────────────────────────────────────────
    ws.freeze_panes=f"{get_column_letter(INFO+1)}6"
    ws.sheet_view.zoomScale=90
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToPage=True
    ws.page_setup.fitToWidth=1; ws.page_setup.paperSize=9  # A4

    os.makedirs(os.path.dirname(OUT),exist_ok=True)
    wb.save(OUT)
    print(f"✅  {OUT}")

if __name__=="__main__":
    build()
